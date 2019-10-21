#Ahmed Waseef
import numpy as np
import pandas as pd
import datetime as dt
from datetime import datetime, date, timedelta, time
import os
import pathlib
from collections import Counter
#import py2exe
import openpyxl
from openpyxl.styles import Alignment
import warnings
from openpyxl.styles import PatternFill, Border, Side, Protection, Font
from openpyxl.worksheet.datavalidation import DataValidation
#from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule
warnings.filterwarnings("ignore")


#%%
'''functions'''

def bustype(buslist, btype):
    if buslist==[]: return []
    global busprop
    output=[]
    for i in buslist:
        if 'temp' in i: continue
        if (busprop[busprop['Bus']==i]['Type'].iloc[0]==btype):
            output.append(i)
    return output

def AssignedRoute(buslist, A1):
    if buslist==[]: return []
    global busprop
    A2=A1[4:]+"-"+A1[:3]
    output=[]
    for i in buslist:
        if 'temp' in i: continue
        if ((busprop[busprop['Bus']==i]['Route1'].iloc[0]==A1) 
        or (busprop[busprop['Bus']==i]['Route1'].iloc[0]==A2)
        or (busprop[busprop['Bus']==i]['Route2'].iloc[0]==A1)
        or (busprop[busprop['Bus']==i]['Route2'].iloc[0]==A2)):
            output.append(i)
    return output

def secondary(buslist, routelist):
    if buslist==[]: return []
    global busprop
    output=[]
    for i in buslist:
        if 'temp' in i: continue
        if ((busprop[busprop['Bus']==i]['Route1'].iloc[0] in routelist) 
        or (busprop[busprop['Bus']==i]['Route1'].iloc[0] in routelist)):
            output.append(i)
    return output

def RouteR(buslist, Dep, Arr):
    if buslist==[]: return []
    global busprop
    output=[]
    for i in buslist:
        if 'temp' in i: continue
        if ((Dep not in busprop[busprop['Bus']==i]['Route1'].iloc[0]) 
        and (Arr in busprop[busprop['Bus']==i]['Route1'].iloc[0])):
            output.append(i)
    return output

def pref_r(buslist, route):
    if buslist==[]: return []
    global busprop
    global preferred_r
    route2=route[4:]+"-"+route[:3]
    output=[]
    for i in buslist:
        if 'temp' in i: continue
        if  i not in preferred_r.keys():
            continue
        if (route in preferred_r[i]) or (route2 in preferred_r[i]):
            output.append(i)
    return output

def updateR(sbuscount,Rating):
    rank, count, previous, buscountrank =0, 0, None, {}
    for key, num in sbuscount:    
        count += 0.5
        if num != previous:
            rank += count
            previous = num
            count = 0
        buscountrank[key] = rank
    Ranking = dict(Counter(buscountrank)+Counter(Rating))
    return Ranking

def updateR2(sbuscount,Rating):
    rank, count, previous, buscountrank = len(buscount), 0, None, {}
    for key, num in sbuscount:    
        count += 0.5
        if num != previous:
            rank -= count
            previous = num
            count = 0
        buscountrank[key] = rank
    Ranking = dict(Counter(buscountrank)+Counter(Rating))
    return Ranking

def busassign(busL,Ranking):
    bus=busL[0]
    busrating=Ranking[bus]
    for j in busL:
        if Ranking[j]>busrating:
            busrating=Ranking[j]
            bus=j
    return bus

def calculate_r(cities,tlist,busloc,bustime,busAvailability,standby):

    r={}
    for city in cities:
        temp={}
        count=0
        for t in tlist:
            count+=1 #a vague approximatioon of latest arrival time of previous period
            if count<=200:
                buslist=[]
                buslist2=[]
                for b in busloc:
                    if b not in busAvailability.keys():
                        continue;
                        
                    if (busloc[b]==city and bustime[b]==t and busAvailability[b]=='Available' 
                        and standby[b]=='Operational'):
                        buslist.append(b)
                    elif (busloc[b]==city and bustime[b]==t and busAvailability[b]=='Available'
                          and standby[b]=='Standby'):
                        buslist2.append(b)
                temp[t]={'buslist':buslist, 'buslist2':buslist2}
            else:
                temp[t]={'buslist':[], 'buslist2':[]}
            r[city]=temp
    return r

#%%
'''Input Files'''
print("Reading Input files.....")
busprop=pd.read_excel('INPUT FILE.xlsx', sheet_name='Buses')
stay=pd.read_excel('INPUT FILE.xlsx', sheet_name='Stay')
bustypes=pd.read_excel('INPUT FILE.xlsx', sheet_name='Type')
mismatchinput=pd.read_excel('INPUT FILE.xlsx', sheet_name='Mismatch')
matchlist=pd.read_excel('MATCHLIST.xlsx', sheet_name='Sheet')
cluster=pd.read_excel('Cluster.xlsx', sheet_name='Cluster')
schedule=pd.read_excel('Schedule.xls', sheet_name='Sheet1')

mmcase=mismatchinput.to_dict()
stay['Stay']=stay['Stay'].apply(lambda x:dt.timedelta(hours=x.hour, minutes=x.minute))

preferred_r={}
for i in range(len(list(cluster['Bus']))):
    bus=cluster['Bus'].iloc[i]
    preferred_r[bus]=[]
    for j in range(10):
        route=cluster["Route-%s" % str(j)].iloc[i]
        if route!=route:
            break
        preferred_r[bus].append(route)

#%%
'''df calculation'''
print("bus previous history.....")  

df = matchlist.rename(columns={'Bus #': 'Bus','Travel Time':'Trip','Route Travel':'Route11'})
busprop["Bus"]=busprop["Bus"].apply(str)
df["Bus"]=df["Bus"].apply(str)
#give mylist and stay as input and get df as output
df=df[df["Bus"]!="Drop"]
df=df[df["Bus"]!="DROP"]
df=df[df["Route"]!="RWP-MRE"]
df=df[df["Route"]!="MRE-RWP"]


#df["Trip"]=df["Trip"].astype(int)
df.loc[df['Trip']!=df['Trip'],'Trip']=2359 #if no trip time then its a breakdown and we put travel time as 24 hours
df["Trip"]=df["Trip"].apply(int)
df["Trip"]=df["Trip"].apply(str)
df["min"]=df["Trip"].apply(lambda x:x[len(x)-2:]).apply(int)
df["hour"]=df["Trip"].apply(lambda x:x[0:len(x)-2] if len(x)>2 else '0').apply(int)
df["trip"]=(df["min"].apply(lambda x: np.timedelta64(x,'m'))+
  df["hour"].apply(lambda x: np.timedelta64(x,'h')))
df=df.drop(columns=['Trip','min','hour'])

df["From"]=df["From"].apply(lambda x: x.upper())
df["To"]=df["To"].apply(lambda x: x.upper())

df["From"]=df["From"].replace("HYD", "HDC")
df["To"]=df["To"].replace("HYD", "HDC")

df.loc[df["Schedule departure"]!=df["Schedule departure"],"Schedule departure"]=time(0,0)
df["Dep Time"]=(df["Schedule departure"].apply(lambda x: 
    pd.Timedelta(hours=x.hour, minutes=x.minute)))
df["datetime"]=pd.to_datetime(df["Dep Time"]+df["Date"])
df=df.drop(columns=["Schedule departure",'Actual Departure Time'])

df["depdate"]=(df["datetime"].apply(lambda x:x if int(x.strftime("%M"))%15==0 else 
  x+pd.to_timedelta(15-int(x.strftime("%M"))%15,'m')))
df["arrdate"]=df["depdate"]+df["trip"]
df["arrdate"]=(df["arrdate"].apply(lambda x:x if int(x.strftime("%M"))%15==0 else 
  x+pd.to_timedelta(15-int(x.strftime("%M"))%15,'m')))
df.loc[df["Terminal"]=="BD","arrdate"]=df["arrdate"]+np.timedelta64(3,'h')
df['triptime']=df['arrdate']-df['depdate']
df['Departure']=df['Route'].apply(lambda x:x[:3])

df["Departure"]=df["Departure"].apply(lambda x: x.upper())
df["Departure"]=df["Departure"].replace("HYD", "HDC")

df=pd.merge(df, stay, on='Route11', how='left')
df.loc[df['Stay']!=df['Stay'],'Stay']=pd.Timedelta('0 days 01:30:00')
df=df.reset_index()
#get output here

#%%
''' INPUT Schedule Provided by the Operations Department '''
print("Reading Schedule to plan.....")

sl=schedule[schedule["Status"]!="DROP"]
sl=sl[sl["Bus Type"]!="APV"]
sl=(sl.drop(columns=['Late (Mins)','Create User','Driver ID','Hostess ID','Hostess Name',
                     'Bus Meter Reading','Meter Out','Driver Name','Remarks','Status']))

sl["Trip"]=sl["Travel Time"].apply(str)
sl["min"]=sl["Trip"].apply(lambda x:x[len(x)-2:]).apply(int)
sl["hour"]=sl["Trip"].apply(lambda x:x[0:len(x)-2] if len(x)>2 else '0').apply(int)
sl["trip"]=(sl["min"].apply(lambda x: np.timedelta64(x,'m'))+
  sl["hour"].apply(lambda x: np.timedelta64(x,'h')))
sl["t"] = sl["Scheduled Departure Time"].apply(lambda x:datetime.strptime(x, "%H:%M"))
sl["Dep Time"]=sl["t"].apply(lambda x:pd.Timedelta(hours=x.hour, minutes=x.minute))
sl["datetime"]=pd.to_datetime(sl['Date'])+sl["Dep Time"]

sl=sl.drop(columns=['Trip','min','hour','t'])

sl["depdate"]=(sl["datetime"].apply(lambda x:x if int(x.strftime("%M"))%15==0 
  else x+pd.to_timedelta(15-int(x.strftime("%M"))%15,'m')))
sl["actarr"]=sl["datetime"]+sl["trip"]
sl["arrdate"]=sl["depdate"]+sl["trip"]
sl["arrdate"]=(sl["arrdate"].apply(lambda x:x if int(x.strftime("%M"))%15==0 else 
  x+pd.to_timedelta(15-int(x.strftime("%M"))%15,'m')))
sl['triptime']=sl['arrdate']-sl['depdate']
sl['Departure']=sl['Route'].apply(lambda x:x[:3])
sl["Route11"]=sl["Route"].apply(lambda x:x[:3]+"-"+x[-3:])
sl['Arrival']=sl['Route11'].apply(lambda x:x[4:])

sl["Departure"]=sl["Departure"].replace("HYD", "HDC")
sl['Arrival']=sl['Arrival'].replace("HYD", "HDC")

sl=pd.merge(sl, stay, on='Route11', how='left')
sl.loc[sl['Stay']!=sl['Stay'],'Stay']=pd.Timedelta('0 days 01:30:00')
sl=pd.merge(sl, bustypes, on='Bus Type', how='left')

#%%
'''initializations variables and dictionaries calculations'''

print("routes and bus locations calculation.....")

import time
start = time. time()

planfrom=min(list(sl['depdate']))
plantill=max(list(sl['depdate']))
busloc, bustime, busAvailability, standby, buscount, Ranking, Rating, Type, Route1={},{},{},{},{},{},{},{},{}
last_r, last_a, busstay={},{},{}
dfhalf=df[df['depdate']<planfrom]

busprop['rate'] = (((busprop.Breakdown.rank(ascending=0,method='min')*0.15))+
       ((busprop.Year.rank(ascending=1,method='min')*0.25))+
       ((busprop.Accident.rank(ascending=0,method='min')*0.1)))

for i in list(busprop['Bus']):
    tempdf=dfhalf[dfhalf['Bus']==i]
    if len(tempdf)==0:
        continue;
    temproute=tempdf[tempdf.index==tempdf.index.max()]['Route11'].iloc[0]
    if temproute not in list(stay['Route11']):
        tempstay=pd.Timedelta('0 days 01:30:00')
    else:
        tempstay=stay[stay['Route11']==temproute]['Stay'].iloc[0]
    busloc[i]=tempdf[tempdf.index==tempdf.index.max()]['To'].iloc[0]
    last_r[i]=tempdf[tempdf.index==tempdf.index.max()]['Route11'].iloc[0]
    bustime[i]=tempdf[tempdf.index==tempdf.index.max()]['arrdate'].iloc[0]+tempstay
    last_a[i]=tempdf[tempdf.index==tempdf.index.max()]['arrdate'].iloc[0]
    busstay[i]=tempstay
    
    if bustime[i]<planfrom:
        bustime[i]=planfrom-pd.Timedelta('0 days 01:00:00')
    
    busAvailability[i]=busprop[busprop["Bus"]==i]['Availability'].iloc[0]
    standby[i]=busprop[busprop["Bus"]==i]['Standby'].iloc[0]
    Rating[i]=busprop[busprop["Bus"]==i]['rate'].iloc[0]
    Type[i]=busprop[busprop["Bus"]==i]['Type'].iloc[0]
    Route1[i]=busprop[busprop["Bus"]==i]['Route1'].iloc[0]
    buscount[i]=0
    
sbuscount = sorted(buscount.items(), key=lambda item: item[1])
Ranking=updateR(sbuscount,Rating)

i=planfrom
tlist=[planfrom-pd.Timedelta('0 days 01:00:00')]
while i <= plantill:
    tlist.append(i)
    i+=pd.Timedelta('0 days 00:15:00')
tlist.append(i)
cities=list(set(list(df['To'])+list(df['From'])))

a=df.groupby(['Route11']).count()['To']
routeused=pd.DataFrame({'Route':a.index,'Count':a.values})
routeused=routeused[routeused['Count']>5]

rd={}
for city in cities:
    rd[city]={'primaryr':[], 'primaryc':[], 'secondaryr':[]}
    for route in list(routeused['Route']):
        tick=0
        dep=route[:3]
        arr=route[4:]
        if city==dep:
            rd[city]['primaryr'].append(route)
            if arr not in rd[city]['primaryc']:
                pcity=arr
                rd[city]['primaryc'].append(arr)
                tick=1
                
        elif city==arr:
            rd[city]['primaryr'].append(route)
            if dep not in rd[city]['primaryc']:
                pcity=dep
                rd[city]['primaryc'].append(dep)
                tick=1
        if tick==1:
            for route2 in list(set(list(df['Route11']))):
                if pcity in route2 and route2 not in rd[city]['primaryr']:
                    rd[city]['secondaryr'].append(route2)

dfdict=(sl[['Terminal','Departure','Arrival','depdate','triptime','arrdate','Type',
            'Route11','Stay', 'datetime','trip','actarr','Bus Type']].to_dict('split'))
    
df1=sl[['Terminal','Departure','Arrival','depdate','triptime','arrdate','Type',
            'Route11','Stay', 'datetime','trip','actarr','Bus Type']]
co=['Terminal','Departure','Arrival','depdate','triptime','arrdate','Type',
            'Route11','Stay', 'datetime','trip','actarr','Bus Type','Bus','Text','Stay Time']

dfpd1=pd.DataFrame(columns=co)
dfnew=pd.DataFrame(columns=list(df1))
#%%
''' bus assignment'''

print("Assigning buses to schedule.....")

opsbuses=[]
count=0
mismatchcount=0
routecount=0
yearmatchcount=0
lessstay=0

for run in [0,1]:
    
    r=calculate_r(cities,tlist,busloc,bustime,busAvailability,standby)

    if run==0:
        dfcurr=df1
    if run==1:
        dfcurr=dfnew
    
    for t in range(len(tlist)):
        try:
            df2=dfcurr[dfcurr['depdate']==tlist[t]]
        except:
            continue
        
        for s in range(len(df2)):
            text=""
            bus='empty'

            df3=df2.iloc[s]
            
            if run==0:
                limit=1
            else:
                limit=int((df3['Stay'].seconds)/60-15) # this is wrong as this is current route stay and not stay after last travel
            
            for y in range(0,limit+1,15):
                
                '''mismatch apply when'''
                mismatchbarrier=45
                if bus=='empty' and y>mismatchbarrier:
                    for j in range(0,y,15):
                        if len(r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=j)]['buslist'])>0:
                            busL1=r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=j)]['buslist']
                            busL=[i for i in busL1 if mmcase[Type[i]][df3['Type']]==1]
                            if len(busL)>0:
                                Ranking = updateR(sbuscount,Rating)
                                bus=busassign(busL,Ranking)
                                r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=j)]['buslist'].remove(bus)
                                text=str(df3['Type'])+"-"+str(Type[bus])+"-mismatch-"+str(j)
                                mismatchcount+=1
                                lessstay=j
                                break;
                            
                if bus!='empty':
                    break;
                
                if bus=='empty' and y>mismatchbarrier:
                    for j in range(0,y,15):  
                        if len(r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=j)]['buslist2'])>0:
                            busL1=r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=j)]['buslist2']
                            busL=[i for i in busL1 if mmcase[Type[i]][df3['Type']]==1]                  
                            if len(busL)>0:
                                Ranking = updateR(sbuscount,Rating)
                                bus=busassign(busL,Ranking)
                                r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=j)]['buslist2'].remove(bus)
                                text=str(df3['Type'])+"-"+str(Type[bus])+"-mismatch-"+str(j)
                                mismatchcount+=1
                                lessstay=j
                                break;
    
                if bus!='empty':
                    break;
                            
                if (len(r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist'])==0 and 
                    len(r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist2'])==0):
                    continue;
    
                usedlist=bustype(r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist'],df3['Type'])
                Routematch=AssignedRoute(usedlist,df3['Route11'])
                Arrmatch=RouteR(usedlist,df3['Departure'],df3['Arrival'])
                prefmatch=pref_r(usedlist,df3['Route11'])
                secondarymatch=secondary(usedlist,rd[df3['Arrival']]['secondaryr'])
                
                usedlistS=bustype(r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist2'],df3['Type'])
                ArrmatchS=RouteR(usedlistS,df3['Departure'],df3['Arrival'])
                secondarymatchS=secondary(usedlistS,rd[df3['Arrival']]['secondaryr'])
                    
                if y==0 and len(Routematch)==0: indicator=0
                else: indicator=1
    
                if indicator==0:
                    increment=0
                    Routematch_temp=Routematch
                    while len(Routematch_temp)==0 and increment<44:
                        increment+=15
                        try:
                            usedlist_temp=bustype(r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=increment)]['buslist'],df3['Type'])
                        except:
                            usedlist_temp=[]
                        Routematch_temp=AssignedRoute(usedlist_temp,df3['Route11']) 
                        
                        if len(Routematch_temp)>0:
                            busL=Routematch_temp
                            Ranking = updateR2(sbuscount,Rating)
                            bus=busassign(busL,Ranking)
                            r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=increment)]['buslist'].remove(bus)
                            text='on route-'+str(increment)+'loop'
                            routecount+=1
                            lessstay=increment
                        if increment>29 and bus=='empty':
                            indicator=1
                            
                if bus!='empty':
                    break;
    
                if indicator==1:
                    if len(Routematch)>0:
                        busL=Routematch
                        Ranking = updateR2(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist'].remove(bus)
                        routecount+=1            
                        text='on route-'+str(y)
                        lessstay=y
                        
                    elif len(Arrmatch)>0 and run==1:
                        busL=Arrmatch
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist'].remove(bus)
                        text='Arrival match-'+str(y)
                        lessstay=y
                        
                    elif len(prefmatch)>0 and run==1:
                        busL=prefmatch
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist'].remove(bus)
                        text='Pref match-'+str(y)
                        lessstay=y
                        
                    elif len(secondarymatch)>0 and run==1:
                        busL=secondarymatch
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist'].remove(bus)
                        text='secondary match-'+str(y) 
                        lessstay=y
        
                    elif len(usedlist)>0 and run==1:
                        busL=usedlist
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist'].remove(bus)
                        text='type match-'+str(y)
                        lessstay=y
                        
                    elif len(ArrmatchS)>0 and run==1:
                        busL=ArrmatchS
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist2'].remove(bus)
                        text='Arrival match Standby-'+str(y)
                        lessstay=y
                        
                    elif len(prefmatch)>0 and run==1:
                        busL=prefmatch
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist2'].remove(bus)
                        text='Pref match- standby'+str(y)
                        lessstay=y
                          
                    elif len(secondarymatchS)>0 and run==1:
                        busL=secondarymatchS
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist2'].remove(bus)
                        text='secondary match- standby'+str(y) 
        
                    elif len(usedlistS)>0 and run==1:
                        busL=usedlistS
                        Ranking = updateR(sbuscount,Rating)
                        bus=busassign(busL,Ranking)
                        r[df3['Departure']][df3['depdate']+pd.Timedelta(minutes=y)]['buslist2'].remove(bus)
                        text='type match- standby'+str(y)
                        lessstay=y
                    
                if bus!='empty':
                    break;
            
            if bus=='empty':
                if run==0:
                    dfnew=dfnew.append(df3)
                    continue;
                elif run==1:
                    bus='temp'+str(count)
                    count+=1
                    text='no bus'
                
            if 'temp' in bus: #if particular temp is being used multiple times(text changes) 
                text='no bus'
            opsbuses.append(bus)
            
            df3['Bus']=bus
            df3['Text']=text
#            if bus in opsbuses+busloc.keys():
            df3['Stay Time']=0
            try:
                df3['Stay Time']=df3['depdate']-last_a[bus]
            except:
                df3['Stay Time']=pd.Timedelta(minutes=121)
            
            last_r[bus]=df3['Route11']
            last_a[bus]=df3['arrdate']
            busstay[bus]=df3['Stay']
            
            dfpd1=dfpd1.append(df3)
            if bus not in Type.keys():
                Type[bus]=0
            if bus in buscount.keys() and 'temp' not in bus:
                buscount[bus]+=1
            else:
                buscount[bus]=0
            sbuscount = sorted(buscount.items(), key=lambda item: item[1])
            
            arr=df3['arrdate']+df3['Stay']
            busloc[bus]=df3['Arrival']
            bustime[bus]=arr
            if arr not in tlist:
                tlist.append(arr)
                tlist=sorted(tlist)
                for city in cities:
                    r[city][arr]={'buslist':[],'buslist2':[]}
    
            if 'temp' in bus:
                r[df3['Arrival']][arr]['buslist']=[bus]+r[df3['Arrival']][arr]['buslist']
            elif busprop[busprop['Bus']==bus]['Standby'].iloc[0]=="Operational":
                r[df3['Arrival']][arr]['buslist']=[bus]+r[df3['Arrival']][arr]['buslist']
            else:
                r[df3['Arrival']][arr]['buslist2']=[bus]+r[df3['Arrival']][arr]['buslist2']

        if t==len(tlist)-1:
            continue
        for city in cities:
            r[city][tlist[t+1]]['buslist']=r[city][tlist[t]]['buslist']+r[city][tlist[t+1]]['buslist']
            r[city][tlist[t+1]]['buslist2']=r[city][tlist[t]]['buslist2']+r[city][tlist[t+1]]['buslist2']

dfpd1=(dfpd1.rename(index=str, columns={'depdate':"Deptime_new",'triptime':"traveltime_new",
                                            'arrdate':"Arrtime_new",'Route11':"Route",
                                            'datetime':"Departure Time",'trip':"Travel Time",
                                            'actarr':"Arrival Time"}))

dfpd=(dfpd1[['Terminal','Route','Departure','Arrival','Departure Time','Travel Time',
            'Arrival Time','Bus','Bus Type','Type','Text','Stay Time']])
    
dfpd['Dep_date'] = dfpd['Departure Time'].apply(lambda x:x.date())
dfpd['Dep_time'] = dfpd['Departure Time'].apply(lambda x:x.time())

dfpd['Trip Time']=(dfpd['Travel Time'].apply(lambda x:
    str(int(24*x.days+(x.seconds-x.seconds%(60*60))/(60*60)))+":"+"0"+str(int(x.seconds%(60*60)/60)) if
    int(x.seconds%(60*60)/60)<10
    else str(int(24*x.days+(x.seconds-x.seconds%(60*60))/(60*60)))+":"+str(int(x.seconds%(60*60)/60))))

dfpd['Stay Time']=(dfpd['Stay Time'].apply(lambda x:
    str(int(24*x.days+(x.seconds-x.seconds%(60*60))/(60*60)))+":"+"0"+str(int(x.seconds%(60*60)/60)) if
    int(x.seconds%(60*60)/60)<10
    else str(int(24*x.days+(x.seconds-x.seconds%(60*60))/(60*60)))+":"+str(int(x.seconds%(60*60)/60))))


dfpd=dfpd.sort_values(by=['Departure Time','Route'])
dfpd.index = np.arange(1,len(dfpd)+1)

#%%
# Output given to the Operation Department
print("preparing output.....")


sb=[i for i in busloc.keys() if i not in opsbuses if busAvailability[i]=='Available']
sby=pd.DataFrame()
for i in sb:
    sby=(sby.append({'Bus':i,'Terminal':busloc[i], 
                     'Route':busprop[busprop['Bus']==i]['Route1'].iloc[0],
                     'Status':standby[i]}, ignore_index=True))
sby=sby[['Bus','Terminal','Route','Status']] #rearranging
sby.index = np.arange(1,len(sby)+1)

nobus=dfpd[dfpd['Text']=='no bus']
nobus.index = np.arange(1,len(nobus)+1)

usage=pd.DataFrame()
usage['Bus']=list(busprop['Bus'])
for i in dfpd['Dep_date'].unique():
    usage[i]=0
'''###work on this in case the schedule plan is for more than one days
'''
a=dfpd.groupby(['Bus']).count()['Departure']
usage=pd.DataFrame({'Bus':a.index,'count':a.values})
usage.index = np.arange(1,len(usage)+1)

dfpd=pd.merge(dfpd, usage, on='Bus', how='left') 

Location=pd.DataFrame.from_dict(busloc,orient='index',columns=['Terminal'])
BTime=pd.DataFrame.from_dict(bustime,orient='index',columns=['Time'])
Location=Location.reset_index()
BTime=BTime.reset_index()
Location = Location.rename(columns={'index': 'Bus'})
BTime = BTime.rename(columns={'index': 'Bus'})
Merge1=pd.merge(Location,BTime[['Bus','Time']],on='Bus', how='left')
BusDetails=pd.merge(usage,Merge1[['Bus','Terminal',"Time"]],on='Bus', how='left')
BusDetails.index = np.arange(1,len(BusDetails)+1)

#%%
b=dfpd['Text'].value_counts()
Summary=pd.DataFrame({'Bus':b.index,'count':b.values})
Summary=Summary.append({'Bus':'Total Buses Used', 'count':(dfpd['Bus'].nunique())}, ignore_index=True)
Summary=Summary.append({'Bus':'Total Depatures', 'count':(len(dfpd))}, ignore_index=True)
Summary=Summary.append({'Bus':'Total No Buses Departures', 'count':(len(nobus))}, ignore_index=True)
Summary.index = np.arange(1,len(Summary)+1)

#%%
today = date.today()
print("exporting solution to excel.....")

datestring = today.strftime('%d-%m-%Y')
pathlib.Path('Output-2' + datestring).mkdir(parents=True)
os.chdir('Output-2' + datestring)
writer = pd.ExcelWriter(datestring + '.xlsx')
(dfpd[['Terminal','Departure','Arrival','Dep_date','Dep_time','Trip Time',
            'Arrival Time','Bus','Bus Type','Type','Text','Stay Time']].to_excel(writer,'Plan'))
(nobus[['Terminal','Departure','Arrival','Dep_date','Dep_time','Trip Time',
            'Arrival Time','Bus','Bus Type','Type','Text']].to_excel(writer,'No Bus Detail'))   
sby.to_excel(writer,'Standby')
usage.to_excel(writer,'Usage')
BusDetails.to_excel(writer,'Bus Details')
Summary.to_excel(writer,'Summary')
writer.save()

#(dfpd[['Departure','Arrival','Dep_date','Dep_time','Trip Time','Arrival Time','Bus','Bus Type']].to_excel(writer,'Plan'))

wb = openpyxl.load_workbook(datestring + '.xlsx')
Plan = wb.active
Plan.column_dimensions['B'].width = 23
Plan.column_dimensions['C'].width = 13
Plan.column_dimensions['E'].width = 13
Plan.column_dimensions['F'].width = 13
Plan.column_dimensions['H'].width = 18
Plan.column_dimensions['L'].width = 23
Plan.column_dimensions['M'].width = 13
Plan.column_dimensions['G'].width = 13
Plan.column_dimensions['I'].width = 12

Plan.freeze_panes = 'A2'
blueFill = PatternFill(start_color='6f9bed', end_color='6f9bed', fill_type='solid')
greenFill = PatternFill(start_color='38c272', end_color='38c272', fill_type='solid')
Plan.conditional_formatting.add('I1:I5000', FormulaRule(formula=['countif($I$2:$I$5000,I1)>1'], stopIfTrue=True, fill=blueFill))
row_count = Plan.max_row
column_count = Plan.max_column
rows = range(1, row_count+1)
columns = range(1, column_count+1)
for row in rows:
    for col in columns:
        Plan.cell(row, col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

sheet2 = wb['No Bus Detail']
sheet2.column_dimensions['B'].width = 23
sheet2.column_dimensions['C'].width = 13
sheet2.column_dimensions['E'].width = 13
sheet2.column_dimensions['F'].width = 13
sheet2.column_dimensions['H'].width = 18
sheet2.column_dimensions['L'].width = 23
sheet2.column_dimensions['M'].width = 13
sheet2.column_dimensions['G'].width = 13
sheet2.column_dimensions['I'].width = 12
sheet2.freeze_panes = 'A2'
row_count = sheet2.max_row
column_count = sheet2.max_column
rows = range(1, row_count+1)
columns = range(1, column_count+1)
for row in rows:
    for col in columns:
        sheet2.cell(row, col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
sheet = wb['Bus Details']
sheet.column_dimensions['E'].width = 18
sheet.freeze_panes = 'A2'
sheet0 = wb['Standby']
sheet0.column_dimensions['E'].width = 13
sheet0.column_dimensions['E'].width = 13
sheet0.freeze_panes = 'A2'
sheet1 = wb['Summary']
sheet1.column_dimensions['B'].width = 25
sheet1.freeze_panes = 'A2'
wb.save(datestring + '.xlsx')

#dfpd.sort_values('Terminal', inplace=True)
#dfpd.set_index(keys=['Terminal'], drop=False,inplace=True)
Terminal=dfpd['Departure'].unique().tolist()
    
dests={}

for city in Terminal:
    file=(dfpd.loc[dfpd.Departure==city][['Departure','Arrival','Dep_date','Dep_time',
          'Trip Time','Arrival Time','Bus','Bus Type','Stay Time','count']])
    file.index = np.arange(1,len(file)+1)
    sbycity=sby[sby['Terminal']==city]
    sbycity.index = np.arange(1,len(sbycity)+1)
    
    arrivals=file['Arrival'].unique().tolist()
    writer = pd.ExcelWriter("%s.xlsx" % city)
    file.to_excel(writer, "%s" % city)
    sbycity.to_excel(writer, 'Standby')
    dests[city]=arrivals
    file0=(dfpd.loc[dfpd.Departure==city][['Departure','Arrival','Dep_date',
           'Dep_time','Bus','Bus Type','Stay Time','count']])
    file0.index = np.arange(1,len(file0)+1)
    file0.to_excel(writer, 'Replies')
    
    for dest in arrivals:
        file1=(file.loc[file.Arrival==dest][['Departure','Arrival','Dep_date','Dep_time',
          'Trip Time','Arrival Time','Bus','Bus Type','Stay Time','count']])
        file1.index = np.arange(1,len(file1)+1)
        file1.to_excel(writer, "%s" % dest)
    writer.save()

for city in Terminal:
    wb = openpyxl.load_workbook("%s.xlsx" % city)
    sheet=wb.active
    sheet.column_dimensions['D'].width = 13
    sheet.column_dimensions['G'].width = 18
    sheet.freeze_panes = 'A2'
    sheet.conditional_formatting.add('H2:H200', FormulaRule(formula=['K2>1'], stopIfTrue=True, fill=greenFill))
    sheet.column_dimensions['K'].hidden= True
    row_count = Plan.max_row
    column_count = Plan.max_column
    rows = range(1, row_count+1)
    columns = range(1, column_count+1)
    for row in rows:
        for col in columns:
            Plan.cell(row, col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for dest in dests[city]:
        sheet = wb[dest]
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['G'].width = 18
        sheet.conditional_formatting.add('H2:H200', FormulaRule(formula=['K2>1'], stopIfTrue=True, fill=greenFill))
        sheet.column_dimensions['K'].hidden= True
        sheet.freeze_panes = 'A2'
        row_count = Plan.max_row
        column_count = Plan.max_column
        rows = range(1, row_count+1)
        columns = range(1, column_count+1)
        for row in rows:
            for col in columns:
                Plan.cell(row, col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        sheet2 = wb['Replies']
        sheet2.conditional_formatting.add('F2:F200', FormulaRule(formula=['I2>1'], stopIfTrue=True, fill=greenFill))
        sheet2.column_dimensions['I'].hidden= True
        sheet2.column_dimensions['D'].width = 13
        sheet2.column_dimensions['K'].width = 35
        sheet2.column_dimensions['L'].width = 35
        sheet2.column_dimensions['M'].width = 15
        sheet2.freeze_panes = 'A2'
        sheet2['J1'] = "Bus Sent"
        sheet2['K1'] = "Reason for Sending Different Bus"
        sheet2['L1'] = "Reason of not sending Assigned Bus"
        sheet2['M1'] = "Remarks"
        sheet2['J1'].font = Font(bold=True, color='FF0000FF')
        sheet2['K1'].font = Font(bold=True, color='FF0000FF')
        sheet2['L1'].font = Font(bold=True, color='FF0000FF')
        sheet2['M1'].font = Font(bold=True, color='FFFF0000')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet2.cell(row=1,column=12).border = thin_border
        sheet2.cell(row=1,column=9).border = thin_border
        sheet2.cell(row=1,column=10).border = thin_border
        sheet2.cell(row=1,column=11).border = thin_border
        dv1 = DataValidation(type="list",formula1='"Late Arrival,Bus Sent Earlier,Docking or Fitness,Breakdown(Unexpected Repairs),Long Repairs,Old Model(Possibility Of Breakdown),Different Route Bus,Bus Type Change,Not Available at the Moment,Replaced a Breakdown Bus,Departure Added,Departure Dropped"', allow_blank=True)
        dv2=(DataValidation(type="list",formula1='"Route Permit Bus Available,Only Bus Available,Same Type Bus Available,New model Bus was Available,Different Bus sent from Other Side,	Departure Added,Departure Dropped"', allow_blank=True))
        sheet2.add_data_validation(dv1)
        sheet2.add_data_validation(dv2)
        row_length=sheet2.max_row
        dv1.add('L2:L%s' %row_length)
        dv2.add('K2:K%s' %row_length)
        wb.save("%s.xlsx" % city)
      
end = time. time()
print ((end - start)/60)
